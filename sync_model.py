"""
Sync Model: Takes snapshot + updated budget → updates Rishca_OS_Financial_Model.xlsx
Injects actual team costs into the P&L and updates assumptions if scenarios provided.
"""

from openpyxl import load_workbook


def update_financial_model(template_path, output_path, snapshot):
    """
    Update Financial Model with actual costs and scenario assumptions.

    P&L sheet cost rows:
      Row 16: People — Founder          (B16:E16 pre-fund quarters, F16:I16 post-fund years)
      Row 17: People — Post-Fund Team   (F17:I17 post-fund years)
      Row 18: Tools & Infrastructure     (B18:E18 pre-fund, F18:I18 post-fund)
      Row 19: Marketing
      Row 20: Legal & Compliance
      Row 21: Office & Misc
      Row 22: Cloud Infrastructure
      Row 24: Total Costs               (=SUM(B16:B22))

    Columns B-E = Pre-fund Q1-Q4 FY26 (quarterly)
    Columns F-I = Post-fund FY27-FY30 (annual)

    Assumptions sheet:
      B30-E30: Founder salary per month (pre-fund quarters)
      B34-B42: Post-fund team roles & monthly salaries
      B43: Total monthly team cost (=SUM(B34:B42))
      B46-B51: Tool costs
      B52: Total monthly tool cost
    """
    wb = load_workbook(template_path)

    team_costs = snapshot.get("team_costs", {})
    assumptions_data = snapshot.get("assumptions", {})
    egp_to_usd = snapshot.get("currency", {}).get("egp_to_usd", 0.020)

    # ========== UPDATE ASSUMPTIONS SHEET ==========
    if "Assumptions" in wb.sheetnames and assumptions_data:
        ws = wb["Assumptions"]

        # Update pricing if provided
        if "pricing_light" in assumptions_data:
            ws["B7"] = assumptions_data["pricing_light"]
        if "pricing_growth" in assumptions_data:
            ws["B8"] = assumptions_data["pricing_growth"]
        if "pricing_pro" in assumptions_data:
            ws["B9"] = assumptions_data["pricing_pro"]
        if "pricing_enterprise" in assumptions_data:
            ws["B10"] = assumptions_data["pricing_enterprise"]

        # Update churn rate
        if "churn_rate" in assumptions_data:
            ws["B27"] = assumptions_data["churn_rate"]

        # Update team salaries from actual data
        # Map actual monthly costs (in USD) to founder salary cells
        annual_usd = team_costs.get("annual_totals_usd", {})

        # If we have actual FY26 quarterly data, update pre-fund founder costs
        monthly_summary = team_costs.get("monthly_summary", [])
        quarterly_costs_usd = _aggregate_quarterly(monthly_summary, egp_to_usd)

        # Update founder monthly salary in Assumptions (B30-E30)
        # These represent monthly salary for each pre-fund quarter
        for i, q_cost in enumerate(quarterly_costs_usd[:4]):
            col = chr(66 + i)  # B, C, D, E
            # Monthly cost = quarterly total / 3
            monthly = q_cost / 3 if q_cost > 0 else None
            if monthly:
                ws[f"{col}30"] = round(monthly, 0)

        # Update post-fund team monthly total if we have projections
        if "fy27" in annual_usd:
            ws["B43"] = round(annual_usd["fy27"] / 12, 0)

    # ========== UPDATE P&L SHEET — ACTUAL COSTS ==========
    if "P&L" in wb.sheetnames:
        ws_pl = wb["P&L"]

        # Pre-fund quarters (B-E): inject actual team costs in USD
        monthly_summary = team_costs.get("monthly_summary", [])
        quarterly_costs_usd = _aggregate_quarterly(monthly_summary, egp_to_usd)

        # Write actual total team costs to People — Founder row (Row 16)
        # These are quarterly totals
        for i, q_cost in enumerate(quarterly_costs_usd[:4]):
            if q_cost > 0:
                col = chr(66 + i)  # B, C, D, E
                ws_pl[f"{col}16"] = round(q_cost, 0)

        # Write actual tool costs if provided
        tool_costs = snapshot.get("tool_costs_quarterly_usd", {})
        if tool_costs:
            for i, key in enumerate(["q1", "q2", "q3", "q4"]):
                if key in tool_costs:
                    col = chr(66 + i)
                    ws_pl[f"{col}18"] = round(tool_costs[key], 0)

    # ========== UPDATE CUSTOMER GROWTH (if scenario changes) ==========
    if "customer_growth" in snapshot and "Customer Waterfall" in wb.sheetnames:
        ws_cw = wb["Customer Waterfall"]
        cg = snapshot["customer_growth"]

        # Pre-fund new customers quarterly (Assumptions B13:E16)
        if "Assumptions" in wb.sheetnames:
            ws_a = wb["Assumptions"]
            prefund = cg.get("prefund_quarterly", [])
            for i, counts in enumerate(prefund[:4]):
                if isinstance(counts, dict):
                    col = chr(66 + i)
                    ws_a[f"{col}13"] = counts.get("light", 0)
                    ws_a[f"{col}14"] = counts.get("growth", 0)
                    ws_a[f"{col}15"] = counts.get("pro", 0)
                    ws_a[f"{col}16"] = counts.get("enterprise", 0)

    wb.save(output_path)
    return output_path


def _aggregate_quarterly(monthly_summary, egp_to_usd):
    """
    Aggregate monthly team costs into FY26 quarters (USD).
    Q1 = Mar-May 2025, Q2 = Jun-Aug 2025, Q3 = Sep-Nov 2025, Q4 = Dec 2025-Feb 2026
    """
    quarters = [0.0, 0.0, 0.0, 0.0]  # Q1, Q2, Q3, Q4

    quarter_map = {
        (2025, 3): 0, (2025, 4): 0, (2025, 5): 0,
        (2025, 6): 1, (2025, 7): 1, (2025, 8): 1,
        (2025, 9): 2, (2025, 10): 2, (2025, 11): 2,
        (2025, 12): 3, (2026, 1): 3, (2026, 2): 3,
    }

    for month_data in monthly_summary:
        parts = month_data["month"].split("-")
        year, month = int(parts[0]), int(parts[1])
        qi = quarter_map.get((year, month))
        if qi is not None:
            total_egp = month_data.get("total_egp", 0)
            quarters[qi] += total_egp * egp_to_usd

    return quarters
