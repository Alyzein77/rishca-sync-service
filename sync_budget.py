"""
Sync Budget: Takes Supabase snapshot → updates Team_Budget_Clean.xlsx
Writes actual time entries and monthly cost aggregations.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, numbers


# Monthly Summary sheet layout:
# Row 1: Header (A=Name, B=Mar-25, C=Apr-25, ... M=Feb-26)
# Rows 2-12: Team members
# Row 13: blank
# Rows 14-20: Tools
# Row 21: Total Tools
# Row 22: Total Team Costs  (=SUM(B2:B12))
# Row 23: Total Monthly Spend (=B22+B21)
# Row 24: Cumulative Spend

# Month column mapping: B=Mar-25(3), C=Apr-25(4), ... M=Feb-26(14→2)
MONTH_TO_COL = {
    (2025, 3): 2, (2025, 4): 3, (2025, 5): 4, (2025, 6): 5,
    (2025, 7): 6, (2025, 8): 7, (2025, 9): 8, (2025, 10): 9,
    (2025, 11): 10, (2025, 12): 11, (2026, 1): 12, (2026, 2): 13,
}

# Team member name → row mapping in Monthly Summary
MEMBER_ROWS = {
    "Rana El Sobky": 2, "Abaza": 3, "Alaa Ashraf": 4,
    "Nada Amin": 5, "Ahmed Hamdy": 6, "Yasseen Nouh": 7,
    "Amr Tarek": 8, "Tarek Mohamed": 9, "Amal Hamdy": 10,
    "Anas Emad": 11, "Bahaa Mohamed": 12,
}

# Name normalization (Vitalis/Supabase names → budget names)
NAME_MAP = {
    "Aml Hamdy": "Amal Hamdy",
    "Bahaa Lashin": "Bahaa Mohamed",
}


def normalize_name(name):
    return NAME_MAP.get(name, name)


def update_budget_xlsx(template_path, output_path, snapshot):
    """Update Team_Budget_Clean.xlsx with data from the snapshot."""
    wb = load_workbook(template_path)

    # --- Update Time Log sheet ---
    if "Time Log" in wb.sheetnames:
        ws_log = wb["Time Log"]
        entries = []
        for month_data in snapshot.get("team_costs", {}).get("monthly_summary", []):
            for member in month_data.get("by_member", []):
                entries.append(member)

        if entries:
            # Find next empty row (column A, starting from row 2)
            next_row = 2
            while ws_log.cell(row=next_row, column=1).value is not None:
                next_row += 1

            for i, entry in enumerate(entries):
                name = normalize_name(entry.get("name", ""))
                year = entry.get("year", 0)
                month = entry.get("month", 0)
                hours = entry.get("hours", 0)
                rate = entry.get("hourly_rate", 0)
                amount = entry.get("amount", 0)

                # Check for duplicate (same name, year, month)
                duplicate = False
                for check_row in range(2, next_row):
                    existing_name = ws_log.cell(row=check_row, column=4).value
                    existing_year = ws_log.cell(row=check_row, column=2).value
                    existing_month = ws_log.cell(row=check_row, column=3).value
                    if existing_name == name and existing_year == year and existing_month == month:
                        # Update existing entry
                        ws_log.cell(row=check_row, column=6, value=rate)
                        ws_log.cell(row=check_row, column=7, value=hours)
                        ws_log.cell(row=check_row, column=8, value=amount)
                        duplicate = True
                        break

                if not duplicate:
                    row = next_row
                    ws_log.cell(row=row, column=1, value=row - 1)  # #
                    ws_log.cell(row=row, column=2, value=year)
                    ws_log.cell(row=row, column=3, value=month)
                    ws_log.cell(row=row, column=4, value=name)
                    ws_log.cell(row=row, column=5, value=entry.get("employment_type", "Hourly"))
                    ws_log.cell(row=row, column=6, value=rate)
                    ws_log.cell(row=row, column=7, value=hours)
                    ws_log.cell(row=row, column=8).value = f"=F{row}*G{row}"
                    ws_log.cell(row=row, column=9, value="Open")
                    next_row += 1

    # --- Update Monthly Summary with aggregated amounts ---
    if "Monthly Summary" in wb.sheetnames:
        ws_summary = wb["Monthly Summary"]

        for month_data in snapshot.get("team_costs", {}).get("monthly_summary", []):
            parts = month_data["month"].split("-")
            year, month = int(parts[0]), int(parts[1])
            col = MONTH_TO_COL.get((year, month))
            if col is None:
                continue

            # Aggregate by member for this month
            member_totals = {}
            for entry in month_data.get("by_member", []):
                name = normalize_name(entry.get("name", ""))
                member_totals[name] = member_totals.get(name, 0) + entry.get("amount", 0)

            for name, total in member_totals.items():
                row = MEMBER_ROWS.get(name)
                if row:
                    ws_summary.cell(row=row, column=col, value=total)

    # --- Update tool costs if provided ---
    if "tool_costs" in snapshot and "Monthly Summary" in wb.sheetnames:
        ws_summary = wb["Monthly Summary"]
        tool_rows = {
            "Figma": 14, "Lovable": 15, "Google": 16,
            "Zoho": 17, "Make": 18, "Canva": 19, "ClickUp": 20,
        }
        for tool_name, cost_by_month in snapshot.get("tool_costs", {}).items():
            row = tool_rows.get(tool_name)
            if not row:
                continue
            for month_key, amount in cost_by_month.items():
                parts = month_key.split("-")
                year, month = int(parts[0]), int(parts[1])
                col = MONTH_TO_COL.get((year, month))
                if col:
                    ws_summary.cell(row=row, column=col, value=amount)

    wb.save(output_path)
    return output_path
